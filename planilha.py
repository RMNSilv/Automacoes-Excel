import openpyxl
from openpyxl.styles import PatternFill, Border,Side,Alignment,Protection,Font
from openpyxl import Workbook,load_workbook
from openpyxl.utils import get_column_letter




wb = load_workbook('Vendas.xlsx')
ws = wb.active


#Definição Cabeçalho
cabecalho = ws['B1:F1']

#Formatação Negrito
font = Font(name='Arial', size=12,bold=True)

#Formatação das Bordas
border = Border(
    left=Side(border_style="thin", color='00000000'),
    right=Side(border_style="thin", color='00000000'),
    top=Side(border_style="thin", color='00000000'),
    bottom=Side(border_style="thin", color='00000000')
)


#Formatação do preenchimento das células
fill = PatternFill(start_color="C0C0C0", fill_type="solid")

# Aplicar formatação às células do cabeçalho
for row in cabecalho:
    for cell in row:
        cell.font = font
        cell.border = border
        cell.fill = fill


# Formatar a coluna A
col_a = ws.column_dimensions['A']
col_a.font = Font(bold=True)
col_a.fill = PatternFill(fill_type='solid', fgColor='C0C0C0')


# Formatar a coluna F
col_f = ws.column_dimensions['F']
col_f.number_format = 'R$ #,##0.00'


# Formatar todas as células com bordas preenchidas
all_cells = ws['A1:F{}'.format(ws.max_row)]
all_border = Border(top=Side(style='medium'), bottom=Side(style='medium'),
                    left=Side(style='medium'), right=Side(style='medium'))

for row in all_cells:
    for cell in row:
        cell.border = all_border

ws.row_dimensions[1].height = 30

for column in cabecalho:
    max_length = 0
    column = [cell for cell in column]
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2) * 1.2
    ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width


wb.save('VendasFormat.xlsx')


