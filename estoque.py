from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.chart import BarChart, Reference

# Criação da planilha
wb = Workbook()
sheet = wb.active
sheet.title = "Controle de Estoque"

# Cabeçalho da planilha
sheet['A1'] = "Produto"
sheet['B1'] = "Quantidade"
sheet['C1'] = "Categoria"

# Dados de exemplo (substitua por seus dados reais)
dados = [
    ("Parafuso 1", 100, "Categoria 1"),
    ("Parafuso 2", 200, "Categoria 1"),
    ("Parafuso 3", 150, "Categoria 2"),
    ("Parafuso 4", 80, "Categoria 3"),
    ("Parafuso 5", 120, "Categoria 3"),
    ("Parafuso 6", 90, "Categoria 4"),
    ("Parafuso 7", 180, "Categoria 5")
]

# Preenchimento dos dados na planilha
for row, (produto, quantidade, categoria) in enumerate(dados, start=2):
    sheet.cell(row=row, column=1).value = produto
    sheet.cell(row=row, column=2).value = quantidade
    sheet.cell(row=row, column=3).value = categoria

# Estilizando o cabeçalho
header_font = Font(bold=True)
for cell in sheet['A1:C1']:
    cell[0].font = header_font

# Criação da sheet com as informações adicionais
sheet2 = wb.create_sheet(title="Informações Adicionais")
sheet2['A1'] = "Valor total do estoque"
sheet2['A2'] = f"=SUM('Controle de Estoque'!B2:B8)"
sheet2['A4'] = "Valor total por categoria"
sheet2['B4'] = "Quantidade no estoque por categoria"

# Cálculo do valor total e quantidade por categoria
categorias = set([dados[i][2] for i in range(len(dados))])
row = 5
for categoria in categorias:
    soma_valor = f"=SUMIFS('Controle de Estoque'!B:B, 'Controle de Estoque'!C:C, \"{categoria}\")"
    soma_quantidade = f"=SUMIFS('Controle de Estoque'!B:B, 'Controle de Estoque'!C:C, \"{categoria}\")"
    sheet2[f"A{row}"] = soma_valor
    sheet2[f"B{row}"] = soma_quantidade
    row += 1

# Criação do gráfico de barras
chart = BarChart()
data = Reference(sheet2, min_col=2, min_row=4, max_row=row-1, max_col=2)
categories = Reference(sheet2, min_col=1, min_row=5, max_row=row-1)
chart.add_data(data, titles_from_data=True)
chart.set_categories(categories)
chart.title = "Quantidade de Produtos por Categoria"
chart.x_axis.title = "Categoria"
chart.y_axis.title = "Quantidade"
chart.style = 10
sheet2.add_chart(chart, "D4")

# Salva o arquivo
wb.save("Controle_de_Estoque.xlsx")

