import openpyxl
import os

# Lista de arquivos de vendas
arquivos_vendas = ['Vendas_jan.xlsx', 'Vendas_fev.xlsx', 'Vendas_mar.xlsx', 'Vendas_abr.xlsx', 'Vendas_mai.xlsx', 'Vendas_jun.xlsx']

# Criar um novo arquivo Excel
wb = openpyxl.Workbook()

# Percorrer os arquivos de vendas e copiar as abas para o novo arquivo
for arquivo in arquivos_vendas:
    # Abrir o arquivo de vendas
    vendas_wb = openpyxl.load_workbook(arquivo)
    
    # Percorrer as abas do arquivo de vendas
    for sheet_name in vendas_wb.sheetnames:
        # Copiar a aba para o novo arquivo
        vendas_sheet = vendas_wb[sheet_name]
        new_sheet = wb.create_sheet(title=sheet_name)
        
        for row in vendas_sheet:
            for cell in row:
                new_sheet[cell.coordinate].value = cell.value
    
    # Fechar o arquivo de vendas
    vendas_wb.close()

# Remover a aba padrão criada pelo Workbook
default_sheet = wb['Sheet']
wb.remove(default_sheet)

# Salvar o novo arquivo consolidado
wb.save('Vendas_agregadas.xlsx')
